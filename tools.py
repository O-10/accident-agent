import os
import json
import tempfile
from google.adk.tools import FunctionTool


def read_pdf(file_path: str) -> str:
    """Lee un archivo PDF y extrae todo su texto.
    
    Args:
        file_path: Ruta completa al archivo PDF.
    
    Returns:
        Texto extraído del PDF.
    """
    try:
        import pdfplumber
    except ImportError:
        return "Error: Instala pdfplumber con: py -m pip install pdfplumber"

    if not os.path.exists(file_path):
        return f"Error: No se encontró el archivo {file_path}"

    try:
        texto_total = []
        with pdfplumber.open(file_path) as pdf:
            for i, pagina in enumerate(pdf.pages, 1):
                texto = pagina.extract_text()
                if texto:
                    texto_total.append(f"--- Página {i} ---\n{texto}")
        
        if not texto_total:
            return "El PDF no contiene texto extraíble (puede ser una imagen escaneada)."
        
        resultado = "\n\n".join(texto_total)
        return f"PDF leído correctamente ({len(pdf.pages)} páginas):\n\n{resultado}"
    except Exception as e:
        return f"Error al leer el PDF: {str(e)}"


def read_excel(file_path: str) -> str:
    """Lee un archivo Excel (.xlsx, .xls) y extrae los datos de todas las hojas.
    
    Args:
        file_path: Ruta completa al archivo Excel.
    
    Returns:
        Contenido del Excel en formato de texto tabulado.
    """
    try:
        import openpyxl
    except ImportError:
        return "Error: Instala openpyxl con: py -m pip install openpyxl"

    if not os.path.exists(file_path):
        return f"Error: No se encontró el archivo {file_path}"

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        resultado = []
        
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            resultado.append(f"\n=== Hoja: {nombre_hoja} ===")
            
            filas = []
            for fila in ws.iter_rows(values_only=True):
                fila_str = [str(c) if c is not None else "" for c in fila]
                filas.append(" | ".join(fila_str))
            
            if filas:
                resultado.append("\n".join(filas[:100]))  # Max 100 filas
                if len(filas) > 100:
                    resultado.append(f"\n... ({len(filas) - 100} filas más)")
            else:
                resultado.append("(hoja vacía)")
        
        return "\n".join(resultado)
    except Exception as e:
        return f"Error al leer el Excel: {str(e)}"


def read_image(file_path: str) -> str:
    """Lee una imagen y describe su contenido usando análisis visual.
    
    Args:
        file_path: Ruta completa a la imagen (JPG, PNG, BMP, GIF).
    
    Returns:
        Información sobre la imagen y su contenido.
    """
    try:
        from PIL import Image
    except ImportError:
        return "Error: Instala Pillow con: py -m pip install Pillow"

    if not os.path.exists(file_path):
        return f"Error: No se encontró el archivo {file_path}"

    try:
        img = Image.open(file_path)
        info = {
            "formato": img.format,
            "modo": img.mode,
            "dimensiones": f"{img.width}x{img.height}",
            "tamaño_bytes": os.path.getsize(file_path),
        }
        
        info_text = " | ".join([f"{k}: {v}" for k, v in info.items()])
        return f"Imagen cargada: {info_text}\n\nPara analizar el contenido visual de la imagen, descríbela en tu mensaje y el agente la interpretará."
    except Exception as e:
        return f"Error al leer la imagen: {str(e)}"


def generate_ishikawa(
    efecto: str,
    mano_obra: str = "",
    metodos: str = "",
    maquinaria: str = "",
    materiales: str = "",
    medio_ambiente: str = "",
    administracion: str = "",
) -> str:
    """Genera un diagrama de Espina de Pesca (Ishikawa) en formato Mermaid.
    
    Args:
        efecto: Efecto o problema a analizar (ej: "Caída desde altura")
        mano_obra: Causas relacionadas con la mano de obra
        metodos: Causas relacionadas con los métodos
        maquinaria: Causas relacionadas con la maquinaria
        materiales: Causas relacionadas con los materiales
        medio_ambiente: Causas relacionadas con el medio ambiente
        administracion: Causas relacionados con la administración
    
    Returns:
        Diagrama en formato Mermaid que puede renderizarse.
    """
    categorias = {
        "Mano de Obra": mano_obra,
        "Métodos": metodos,
        "Maquinaria": maquinaria,
        "Materiales": materiales,
        "Medio Ambiente": medio_ambiente,
        "Administración": administracion,
    }

    lineas = ["graph LR"]
    lineas.append(f"    EFECTO[\"{efecto}\"]")
    
    for i, (cat, causas) in enumerate(categorias.items()):
        if causas:
            nodo_cat = f"C{i}"
            lineas.append(f"    {nodo_cat}[\"{cat}\"]")
            lineas.append(f"    {nodo_cat} --> EFECTO")
            
            lista_causas = [c.strip() for c in causas.split(",") if c.strip()]
            for j, causa in enumerate(lista_causas):
                nodo_causa = f"C{i}_{j}"
                lineas.append(f"    {nodo_causa}(\"{causa}\")")
                lineas.append(f"    {nodo_causa} --> {nodo_cat}")

    diagrama = "\n".join(lineas)
    
    return f"""DIAGRAMA DE ESPINA DE PESCOADO (ISHIKAWA)
==========================================

Efecto: {efecto}

{diagrama}

Para renderizar, copia el código Mermaid y pégalo en:
- https://mermaid.live
- https://www.mermaidchart.com
- O en cualquier visor de Mermaid

---

RESUMEN DE CAUSAS POR CATEGORÍA:
"""
    + "\n".join([f"- {cat}: {c}" for cat, c in categorias.items() if c])


def generate_fta(
    evento_tope: str,
    compuertas: str = "",
) -> str:
    """Genera un Árbol de Fallas (FTA) en formato Mermaid.
    
    Args:
        evento_tope: Evento no deseado (accidente)
        compuertas: Descripción de eventos intermedios y compuertas.
                   Formato: "evento_hijo1|AND|evento_hijo2,evento_hijo3|OR|evento_hijo4"
    
    Returns:
        Diagrama FTA en formato Mermaid.
    """
    lineas = ["graph TD"]
    safe_tope = evento_tope.replace('"', "'")
    lineas.append(f"    TOPE[\"{safe_tope}\"]")
    lineas.append(f"    TOPE --> AND1{{\"AND\"}}")
    
    if compuertas:
        eventos = [e.strip() for e in compuertas.split(",") if e.strip()]
        for i, evento in enumerate(eventos):
            safe_evento = evento.replace('"', "'")
            lineas.append(f"    AND1 --> E{i}[\"{safe_evento}\"]")
    else:
        lineas.append(f"    AND1 --> E0[\"Causa contribuyente 1\"]")
        lineas.append(f"    AND1 --> E1[\"Causa contribuyente 2\"]")

    diagrama = "\n".join(lineas)
    
    return f"""ÁRBOL DE FALLAS (FTA)
====================

Evento Tope: {evento_tope}

{diagrama}

Para renderizar, copia el código Mermaid y pégalo en:
- https://mermaid.live
"""


def generate_bowtie(
    amenaza: str,
    controles_preventivos: str,
    evento: str,
    controles_mitigantes: str,
    consecuencia: str,
) -> str:
    """Genera un diagrama Bow-Tie (Lazo) en formato Mermaid.
    
    Args:
        amenaza: Amenaza o peligro identificado
        controles_preventivos: Controles preventivos (separados por comas)
        evento: Evento central no deseado
        controles_mitigantes: Controles mitigantes (separados por comas)
        consecuencia: Consecuencia del evento
    
    Returns:
        Diagrama Bow-Tie en formato Mermaid.
    """
    safe = lambda s: s.replace('"', "'")
    
    preventivos = [p.strip() for p in controles_preventivos.split(",") if p.strip()]
    mitigantes = [m.strip() for m in controles_mitigantes.split(",") if m.strip()]

    lineas = ["graph LR"]
    lineas.append(f"    AMEN[\"{safe(amenaza)}\"]")
    lineas.append(f"    EVENTO[\"{safe(evento)}\"]")
    lineas.append(f"    CONSE[\"{safe(consecuencia)}\"]")
    
    for i, p in enumerate(preventivos):
        lineas.append(f"    AMEN --> P{i}[\"{safe(p)}\"]")
        lineas.append(f"    P{i} --> EVENTO")
    
    for i, m in enumerate(mitigantes):
        lineas.append(f"    EVENTO --> M{i}[\"{safe(m)}\"]")
        lineas.append(f"    M{i} --> CONSE")

    diagrama = "\n".join(lineas)
    
    return f"""DIAGRAMA BOW-TIE (LAZO)
=====================

Amenaza: {amenaza}
Evento: {evento}
Consecuencia: {consecuencia}

{diagrama}

CONTROLES PREVENTIVOS:
"""
    + "\n".join([f"- {p}" for p in preventivos]) 
    + "\n\nCONTROLES MITIGANTES:\n"
    + "\n".join([f"- {m}" for m in mitigantes])


def generate_5whys(problema: str, niveles: str) -> str:
    """Genera un diagrama de los 5 Porqués en formato Mermaid.
    
    Args:
        problema: Problema o accidente a investigar
        niveles: Niveles de "por qué" en formato: "respuesta1,respuesta2,respuesta3"
    
    Returns:
        Diagrama de 5 Porqués en formato Mermaid.
    """
    respuestas = [r.strip() for r in niveles.split(",") if r.strip()]
    
    lineas = ["graph TD"]
    safe = lambda s: s.replace('"', "'")
    
    lineas.append(f"    P0[\"{safe(problema)}\"]")
    
    for i, resp in enumerate(respuestas):
        nodo_q = f"Q{i+1}"
        nodo_r = f"R{i+1}"
        lineas.append(f"    P{i} --> Q{nodo_q}{{\"¿Por qué?\"}}")
        lineas.append(f"    Q{nodo_q} --> R{nodo_r}[\"{safe(resp)}\"]")
        if i < len(respuestas) - 1:
            lineas.append(f"    R{nodo_r} --> P{i+1}")

    if respuestas:
        lineas.append(f"    R{len(respuestas)} -.-> RA[\"CAUSA RAÍZ\"]")

    diagrama = "\n".join(lineas)
    
    return f"""MÉTODO DE LOS 5 PORQUÉS
=====================

Problema: {problema}

{diagrama}

CADENA CAUSAL:
"""
    + "\n".join([f"{i+1}. ¿Por qué? → {r}" for i, r in enumerate(respuestas)]) 
    + "\n\n→ CAUSA RAÍZ IDENTIFICADA"


tools = [
    FunctionTool(func=read_pdf),
    FunctionTool(func=read_excel),
    FunctionTool(func=read_image),
    FunctionTool(func=generate_ishikawa),
    FunctionTool(func=generate_fta),
    FunctionTool(func=generate_bowtie),
    FunctionTool(func=generate_5whys),
]
